"""
输出模块

负责将计算结果导出为格式化的Excel文件，包含学生排名、班级汇总、单科班级分析和单科班级排名
分析和排名表分别合并为单个工作表，各科目纵向排列，每个科目块前添加科目名称行，块间空一行
使用配色方案设置表头、奇偶行、边框和高亮行

版本：4.6.2
新增：每个科目块前添加科目名称行
修复：处理pandas的<NA>值
"""

import pandas as pd
import numpy as np
import os
import warnings
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import Dict, List, Tuple

# 忽略libpng警告（可能由openpyxl或Pillow内部触发）
warnings.filterwarnings("ignore", message="libpng warning: iCCP: known incorrect sRGB profile")

# 常见科目顺序（可根据需要调整）
COMMON_SUBJECT_ORDER = ['语文', '数学', '英语', '物理', '化学', '生物', '历史', '政治', '地理', '体育']

def export_to_excel(student_df: pd.DataFrame, class_df: pd.DataFrame,
                    subject_details: Dict[str, pd.DataFrame],
                    subject_rankings: Dict[str, pd.DataFrame],
                    output_path: str, passing_score: Dict[str, float] = None, excellent_score: Dict[str, float] = None) -> None:
    """
    将学生排名表、班级汇总表、单科班级分析（纵向堆叠）和单科班级排名（纵向堆叠）导出为Excel文件

    参数:
        student_df: 学生排名表DataFrame
        class_df: 班级汇总表DataFrame
        subject_details: 字典，键为科目名，值为该科目的班级分析表DataFrame
        subject_rankings: 字典，键为科目名，值为该科目的班级排名表DataFrame
        output_path: 输出文件路径
    """
    # 确保输出目录存在
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    # 处理DataFrame中的<NA>值，替换为np.nan，以便openpyxl能处理
    student_df = student_df.replace(pd.NA, np.nan)
    class_df = class_df.replace(pd.NA, np.nan)

    # 使用openpyxl引擎写入Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 先创建所有工作表
        student_sheet_name = '学生排名'
        class_sheet_name = '班级汇总'
        analysis_sheet_name = '单科班级分析'
        ranking_sheet_name = '单科班级排名'
        
        # 写入基本表数据
        student_df.to_excel(writer, sheet_name=student_sheet_name, index=False)
        class_df.to_excel(writer, sheet_name=class_sheet_name, index=False)

        # 获取工作簿对象
        workbook = writer.book
        
        # 2. 处理学生排名表
        student_sheet = workbook[student_sheet_name]
        _format_student_sheet(student_sheet, passing_score, excellent_score)
        
        # 3. 处理班级汇总表
        class_sheet = workbook[class_sheet_name]
        _format_class_sheet(class_sheet, class_df)
        
        # 4. 创建并处理单科班级分析工作表
        if subject_details:
            analysis_sheet = workbook.create_sheet(analysis_sheet_name, len(workbook.sheetnames))
            analysis_blocks = _write_blocks(analysis_sheet, subject_details, is_analysis=True)
            _format_block_sheet(analysis_sheet, analysis_blocks, is_analysis=True)
        
        # 5. 创建并处理单科班级排名工作表
        if subject_rankings:
            ranking_sheet = workbook.create_sheet(ranking_sheet_name, len(workbook.sheetnames))
            ranking_blocks = _write_blocks(ranking_sheet, subject_rankings, is_analysis=False)
            _format_block_sheet(ranking_sheet, ranking_blocks, is_analysis=False)
        
        # 6. 为每个班级创建班级详情工作表
        if not student_df.empty:
            # 按照班级数字顺序排序
            import re
            
            def get_class_number(cls):
                match = re.search(r'(\d+)', str(cls))
                if match:
                    return int(match.group(1))
                return 999999
            
            classes = sorted(student_df['班级'].unique(), key=get_class_number)
            for class_name in classes:
                class_sheet_name = f'班级详情_{class_name}'
                class_df = student_df[student_df['班级'] == class_name].copy()
                class_df.to_excel(writer, sheet_name=class_sheet_name, index=False)
                
                # 处理班级详情表
                if class_sheet_name in workbook.sheetnames:
                    class_detail_sheet = workbook[class_sheet_name]
                    _format_student_sheet(class_detail_sheet, passing_score, excellent_score)
        
        # 7. 调整所有工作表的列宽
        for sheet_name in workbook.sheetnames:
            _adjust_column_width(workbook[sheet_name])


def _create_info_sheet(sheet, exam_info):
    """
    创建考试信息首页
    
    参数:
        sheet: openpyxl工作表对象
        exam_info: 字典，包含考试相关信息
    """
    # 设置页面背景色为浅黄色
    sheet.sheet_properties.tabColor = "FFFFCC"
    
    # 设置整个页面的背景色为浅黄色
    for row in range(1, 21):
        for col in range(1, 11):
            cell = sheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    # 设置学年信息
    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=10)
    year_cell = sheet.cell(row=5, column=1)
    year_cell.value = exam_info.get('考试名称', '').split(' ')[0] if exam_info.get('考试名称', '') else ''
    year_cell.font = Font(bold=True, size=16)
    year_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置学校和年段信息
    sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=10)
    school_cell = sheet.cell(row=7, column=1)
    school_cell.value = exam_info.get('考试名称', '').split(' ')[1] if len(exam_info.get('考试名称', '').split(' ')) > 1 else ''
    school_cell.font = Font(bold=True, size=16)
    school_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置表格标题
    sheet.merge_cells(start_row=9, start_column=1, end_row=9, end_column=10)
    table_title_cell = sheet.cell(row=9, column=1)
    table_title_cell.value = exam_info.get('考试名称', '').split(' ')[2] if len(exam_info.get('考试名称', '').split(' ')) > 2 else '质量分析表'
    table_title_cell.font = Font(bold=True, size=20)
    table_title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置内部资料提示
    sheet.merge_cells(start_row=11, start_column=1, end_row=11, end_column=10)
    note_cell = sheet.cell(row=11, column=1)
    note_cell.value = "(内部资料注意保存)"
    note_cell.font = Font(size=12)
    note_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置落款
    sheet.merge_cells(start_row=17, start_column=1, end_row=17, end_column=10)
    signature_cell = sheet.cell(row=17, column=1)
    signature_cell.value = exam_info.get('出卷人', 'xx中学教务处')
    signature_cell.font = Font(size=14)
    signature_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置日期
    sheet.merge_cells(start_row=19, start_column=1, end_row=19, end_column=10)
    date_cell = sheet.cell(row=19, column=1)
    date_cell.value = exam_info.get('考试时间日期', '')
    date_cell.font = Font(size=14)
    date_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    for col in range(1, 11):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    
    # 调整行高
    for row in range(1, 21):
        sheet.row_dimensions[row].height = 25


def _add_sheet_title(sheet, exam_info):
    """
    为工作表添加标题
    
    参数:
        sheet: openpyxl工作表对象
        exam_info: 字典，包含考试相关信息
    """
    if not exam_info or '考试名称' not in exam_info:
        return
    
    # 插入标题行
    sheet.insert_rows(1, 3)
    
    # 合并标题区域
    sheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=10)
    
    # 设置标题背景色
    for row in range(1, 4):
        for col in range(1, 11):
            cell = sheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    # 设置标题文本 - 只对合并单元格的左上角单元格（第1行第1列）进行操作
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = exam_info['考试名称']
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整标题行高
    for row in range(1, 4):
        sheet.row_dimensions[row].height = 20

def _sort_subjects(subjects: List[str]) -> List[str]:
    """按常见科目顺序排序，未知科目放在后面"""
    def key_func(s):
        try:
            return COMMON_SUBJECT_ORDER.index(s)
        except ValueError:
            return len(COMMON_SUBJECT_ORDER)  # 未知科目排在最后
    return sorted(subjects, key=key_func)

def _write_blocks(worksheet, subject_dict: Dict[str, pd.DataFrame], is_analysis: bool) -> List[Tuple[str, int, int]]:
    """
    将多个科目的DataFrame纵向写入工作表，每个块前添加科目名称行，块间空一行

    参数:
        worksheet: openpyxl工作表对象
        subject_dict: 科目名到DataFrame的映射
        is_analysis: 是否分析表（仅用于标识，此处未使用）

    返回:
        blocks: 列表，每个元素为 (科目名, 表头行号, 数据最后一行行号)
    """
    subjects_sorted = _sort_subjects(list(subject_dict.keys()))
    current_row = 1  # 从第1行开始
    blocks = []

    for subject in subjects_sorted:
        df = subject_dict[subject]

        # 如果不是第一个块，先插入一个空行
        if current_row > 1:
            current_row += 1

        # 写入科目名称行（第一列写科目名，其余单元格留空）
        worksheet.cell(row=current_row, column=1, value=subject)
        title_row = current_row
        current_row += 1

        # 写入表头（列名）
        header_row = current_row
        for col_idx, col_name in enumerate(df.columns, start=1):
            try:
                cell = worksheet.cell(row=header_row, column=col_idx)
                cell.value = col_name
            except AttributeError:
                # 忽略合并单元格的只读属性错误
                pass

        # 写入数据行
        data_start_row = header_row + 1
        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=data_start_row):
            for col_idx, value in enumerate(row_data, start=1):
                try:
                    if pd.isna(value):
                        cell_value = None
                    else:
                        cell_value = value
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = cell_value
                except AttributeError:
                    # 忽略合并单元格的只读属性错误
                    pass

        data_end_row = data_start_row + len(df) - 1
        # 记录此块信息
        blocks.append((subject, header_row, data_end_row))

        # 更新当前行为下一个块的起始行
        current_row = data_end_row + 1

    return blocks

def _format_block_sheet(worksheet, blocks: List[Tuple[str, int, int]], is_analysis: bool) -> None:
    """
    对纵向堆叠的块工作表应用样式：每个块内表头特殊，数据行奇偶交替，边框，科目名称行特殊样式

    参数:
        worksheet: 已写入数据的工作表
        blocks: 块信息列表，每个元素为 (科目名, 表头行号, 数据最后一行行号)
        is_analysis: 是否分析表（用于确定百分比列和分数段列）
    """
    from config import COLOR_SUBJECT as color, COLOR_ORANGE_SCALE

    # 边框
    border = _get_border(color['border'])
    # 表头样式
    header_fill = _get_fill(color['header_bg'])
    header_font = Font(bold=True, color=color['header_fg'][1:])
    header_alignment = Alignment(horizontal='center', vertical='center')
    # 数据行样式
    odd_fill = _get_fill(color['odd_row_bg'])
    even_fill = _get_fill(color['even_row_bg'])
    data_alignment = Alignment(horizontal='center', vertical='center')
    # 科目名称行样式（加粗，不同背景色，居中）
    title_fill = _get_fill(color['header_bg'])  # 与表头相同背景，或可自定义
    title_font = Font(bold=True, color=color['header_fg'][1:])
    title_alignment = Alignment(horizontal='left', vertical='center')

    # 百分比列名称（用于分析和排名表）
    pct_columns = ['及格率', '优生率', '平均分比差', '及格率比差', '优生率比差']

    # 不需要额外的行号偏移，因为_write_blocks函数返回的行号已经考虑了标题行的位置

    for subject, header_row, data_end_row in blocks:
        title_row = header_row - 1
        # 直接使用blocks中返回的行号，不需要再偏移
        adjusted_title_row = title_row
        adjusted_header_row = header_row
        adjusted_data_end_row = data_end_row
        
        # 格式化科目名称行
        for col_idx in range(1, worksheet.max_column + 1):
            try:
                cell = worksheet.cell(row=adjusted_title_row, column=col_idx)
                cell.fill = title_fill
                cell.font = title_font
                cell.alignment = title_alignment
                cell.border = border
            except AttributeError:
                # 忽略合并单元格的只读属性错误
                pass

        # 格式化表头行并收集表头信息
        headers = []
        for col_idx in range(1, worksheet.max_column + 1):
            try:
                cell = worksheet.cell(row=adjusted_header_row, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                headers.append(cell.value)
            except AttributeError:
                # 忽略合并单元格的只读属性错误
                headers.append(None)

        # 识别分数段列（仅在分析表中）
        score_segment_columns = []
        class_max_value = 0  # 班级区最大值
        year_max_value = 0   # 年段区最大值
        class_rows = []      # 班级行索引
        year_rows = []       # 年段行索引
        
        if is_analysis:
            score_segment_columns = _get_score_segment_columns(headers)
            
            if score_segment_columns:
                data_start_row = adjusted_header_row + 1
                
                # 首先识别班级行和年段行
                class_column_idx = None
                # 查找班级列
                for i, header in enumerate(headers):
                    if header and str(header).strip() == '班级':
                        class_column_idx = i + 1  # 转换为1-based索引
                        break
                
                # 收集班级行和年段行
                for row_idx in range(data_start_row, adjusted_data_end_row + 1):
                    is_class_row = True
                    if class_column_idx:
                        try:
                            cell = worksheet.cell(row=row_idx, column=class_column_idx)
                            if cell.value:
                                class_name = str(cell.value).strip()
                                # 识别年段行（根据班级名称判断）
                                if class_name in ['年段', '全校', '全年级', '年级', '总']:
                                    is_class_row = False
                                    year_rows.append(row_idx)
                                else:
                                    class_rows.append(row_idx)
                            else:
                                class_rows.append(row_idx)
                        except AttributeError:
                            class_rows.append(row_idx)
                    else:
                        # 如果找不到班级列，默认所有行都是班级行
                        class_rows.append(row_idx)
                
                # 计算班级区的最大值
                class_score_values = []
                for row_idx in class_rows:
                    for col_idx in score_segment_columns:
                        try:
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            if cell.value is not None:
                                try:
                                    val = float(cell.value)
                                    class_score_values.append(val)
                                except (ValueError, TypeError):
                                    pass
                        except AttributeError:
                            pass
                if class_score_values:
                    class_max_value = max(class_score_values)
                
                # 计算年段区的最大值
                year_score_values = []
                for row_idx in year_rows:
                    for col_idx in score_segment_columns:
                        try:
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            if cell.value is not None:
                                try:
                                    val = float(cell.value)
                                    year_score_values.append(val)
                                except (ValueError, TypeError):
                                    pass
                        except AttributeError:
                            pass
                if year_score_values:
                    year_max_value = max(year_score_values)
                
                # 如果班级区或年段区没有值，使用另一个区的最大值作为备用
                if class_max_value == 0 and year_max_value > 0:
                    class_max_value = year_max_value
                elif year_max_value == 0 and class_max_value > 0:
                    year_max_value = class_max_value

        # 格式化数据行（奇偶背景）
        data_start_row = adjusted_header_row + 1
        for row_idx in range(data_start_row, adjusted_data_end_row + 1):
            row_fill = even_fill if (row_idx - data_start_row) % 2 == 0 else odd_fill
            for col_idx in range(1, worksheet.max_column + 1):
                try:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    
                    # 检查是否是分数段列（仅在分析表中）
                    if is_analysis and col_idx in score_segment_columns and cell.value is not None:
                        # 对分数段列应用上色
                        try:
                            value = float(cell.value)
                            # 根据行类型选择对应的最大值
                            if row_idx in class_rows:
                                # 班级行使用班级区的最大值
                                cell.fill = _get_score_segment_color(value, class_max_value, COLOR_ORANGE_SCALE)
                            elif row_idx in year_rows:
                                # 年段行使用年段区的最大值
                                cell.fill = _get_score_segment_color(value, year_max_value, COLOR_ORANGE_SCALE)
                            else:
                                # 其他行使用默认填充
                                cell.fill = row_fill
                        except (ValueError, TypeError):
                            # 如果不是数值，使用默认填充
                            cell.fill = row_fill
                    else:
                        # 非分数段列使用默认填充
                        cell.fill = row_fill
                    
                    cell.alignment = data_alignment
                    cell.border = border

                    # 对百分比列应用数字格式
                    try:
                        col_name = worksheet.cell(row=adjusted_header_row, column=col_idx).value
                        if col_name in pct_columns and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00%'
                    except AttributeError:
                        # 忽略合并单元格的只读属性错误
                        pass
                except AttributeError:
                    # 忽略合并单元格的只读属性错误
                    pass

def _get_fill(color_hex: str) -> PatternFill:
    """根据六位十六进制颜色代码创建填充对象（自动去掉#，并确保为6位）"""
    if color_hex.startswith('#'):
        color_hex = color_hex[1:]
    if len(color_hex) == 6:
        return PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
    else:
        return PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

def _get_border(color_hex: str) -> Border:
    """根据边框颜色创建边框对象"""
    if color_hex.startswith('#'):
        color_hex = color_hex[1:]
    side = Side(style='thin', color=color_hex)
    return Border(left=side, right=side, top=side, bottom=side)


def _get_score_segment_color(value, max_value, color_scale):
    """
    根据分数段人数值获取对应的颜色
    
    参数:
        value: 当前分数段人数
        max_value: 所有分数段人数的最大值
        color_scale: 颜色渐变列表，从浅到深
    
    返回:
        PatternFill对象: 对应的填充颜色
    """
    if max_value <= 1:
        # 如果最大值小于等于1，直接返回最浅的颜色
        return _get_fill(color_scale[0])
    
    # 计算分数区间
    max_value_minus_1 = max_value - 1
    interval = max_value_minus_1 / 4
    
    # 确定颜色索引
    if value == max_value:
        # 最大值使用最深的颜色
        color_index = len(color_scale) - 1
    elif value >= max_value_minus_1 - interval:
        color_index = 3
    elif value >= max_value_minus_1 - 2 * interval:
        color_index = 2
    elif value >= max_value_minus_1 - 3 * interval:
        color_index = 1
    else:
        color_index = 0
    
    return _get_fill(color_scale[color_index])


def _get_score_segment_columns(headers):
    """
    从表头中识别分数段列

    参数:
        headers: 表头列表

    返回:
        list: 分数段列索引列表
    """
    import re
    score_segment_columns = []
    for i, header in enumerate(headers):
        if header:
            # 尝试多种格式匹配分数段
            header_str = str(header).strip()
            # 匹配标准分数段格式，如 "0-10", "10-20" 等
            if re.match(r'^\d+-\d+$', header_str):
                score_segment_columns.append(i + 1)  # 转换为1-based索引
            # 匹配可能的其他格式，如 "0-10分", "10-20分" 等
            elif re.match(r'^\d+-\d+[分%]$', header_str):
                score_segment_columns.append(i + 1)  # 转换为1-based索引
    return score_segment_columns

def _format_student_sheet(worksheet, passing_score: Dict[str, float] = None, excellent_score: Dict[str, float] = None):
    """设置学生排名表的格式：表头颜色、奇偶行颜色、边框、第一名高亮、科目成绩染色"""
    from config import COLOR_STUDENT as color

    header_fill = _get_fill(color['header_bg'])
    header_font = Font(bold=True, color=color['header_fg'][1:])
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = _get_border(color['border'])

    # 表头现在在第1行（因为不再插入标题行）
    header_row = 1
    
    # 添加冻结窗格，冻结第一行
    worksheet.freeze_panes = worksheet['A2']
    
    # 格式化表头行
    for col_idx in range(1, worksheet.max_column + 1):
        try:
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        except AttributeError:
            # 忽略合并单元格的只读属性错误
            pass

    odd_fill = _get_fill(color['odd_row_bg'])
    even_fill = _get_fill(color['even_row_bg'])
    highlight_fill = _get_fill(color['highlight_bg'])
    data_alignment = Alignment(horizontal='center', vertical='center')

    # 从配置文件导入染色颜色方案
    from config import COLOR_SCORE_RANGE
    
    # 定义染色颜色
    excellent_color = _get_fill(COLOR_SCORE_RANGE['excellent'])  # 中蓝
    passing_color = _get_fill(COLOR_SCORE_RANGE['passing'])     # 浅天蓝
    failing_color = _get_fill(COLOR_SCORE_RANGE['failing'])     # 浅蓝
    highest_score_color = _get_fill(COLOR_SCORE_RANGE['highest'])  # 中橙

    # 获取表头，处理可能的合并单元格
    header = []
    for col_idx in range(1, worksheet.max_column + 1):
        try:
            cell = worksheet.cell(row=header_row, column=col_idx)
            header.append(cell.value)
        except AttributeError:
            # 忽略合并单元格的只读属性错误
            header.append(None)
    rank_col_idx = None
    if '年级排名' in header:
        rank_col_idx = header.index('年级排名') + 1

    # 收集所有科目的最高分（包括总分）
    subject_columns = []
    subject_highest_scores = {}
    for col_idx, col_name in enumerate(header):
        if col_name and isinstance(col_name, str) and col_name not in ['班级', '座号', '姓名', '年级排名', '班级排名', '语数英总分', '语数英排名', '上次排名', '进退步']:
            subject_columns.append((col_idx + 1, col_name))  # (列索引, 科目名)
            # 计算该科目的最高分
            highest_score = -1
            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                try:
                    cell = worksheet.cell(row=row_idx, column=col_idx + 1)
                    if isinstance(cell.value, (int, float)) and cell.value > highest_score:
                        highest_score = cell.value
                except AttributeError:
                    pass
            if highest_score >= 0:
                subject_highest_scores[col_name] = highest_score

    # 格式化数据行
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        row_fill = even_fill if (row_idx - header_row - 1) % 2 == 0 else odd_fill
        is_highlight = False
        if rank_col_idx:
            try:
                cell = worksheet.cell(row=row_idx, column=rank_col_idx)
                if isinstance(cell.value, (int, float)) and cell.value == 1:
                    is_highlight = True
            except AttributeError:
                # 忽略合并单元格的只读属性错误
                pass

        # 定义额外的染色颜色
        light_cyan_fill = _get_fill('#b7dee8')    # 浅青色
        sky_blue_fill = _get_fill('#92cddc')      # 天蓝色
        grass_yellow_fill = _get_fill('#d8e4bc')  # 草黄色
        grass_green_fill = _get_fill('#c4d79b')   # 草绿色
        light_pink_purple_fill = _get_fill('#e4dfec')  # 淡粉紫
        purple_fill = _get_fill('#ccc0da')        # 紫色
        
        for col_idx in range(1, worksheet.max_column + 1):
            try:
                cell = worksheet.cell(row=row_idx, column=col_idx)
                col_name = header[col_idx - 1]
                
                # 检查是否需要科目成绩染色（包括总分）
                if col_name and isinstance(col_name, str) and col_name not in ['班级', '座号', '姓名', '年级排名', '班级排名', '语数英总分', '语数英排名', '上次排名', '进退步']:
                    # 检查是否是该科目的最高分
                    if subject_highest_scores.get(col_name, -1) == cell.value:
                        cell.fill = highest_score_color
                    # 否则根据分数范围染色
                    elif isinstance(cell.value, (int, float)):
                        score = cell.value
                        # 获取该科目的阈值
                        passing = passing_score.get(col_name, 60) if passing_score else 60
                        excellent = excellent_score.get(col_name, 80) if excellent_score else 80
                        # 根据分数范围染色
                        if score >= excellent:
                            cell.fill = excellent_color
                        elif score >= passing:
                            cell.fill = passing_color
                        else:
                            cell.fill = failing_color
                    else:
                        cell.fill = row_fill
                # 年段排名用浅青色
                elif col_name and isinstance(col_name, str) and col_name == '年级排名':
                    cell.fill = light_cyan_fill
                # 班级排名用天蓝色
                elif col_name and isinstance(col_name, str) and col_name == '班级排名':
                    cell.fill = sky_blue_fill
                # 语数英总分用草黄色
                elif col_name and isinstance(col_name, str) and col_name == '语数英总分':
                    cell.fill = grass_yellow_fill
                # 语数英排名用草绿色
                elif col_name and isinstance(col_name, str) and col_name == '语数英排名':
                    cell.fill = grass_green_fill
                # 上次排名用淡粉紫
                elif col_name and isinstance(col_name, str) and col_name == '上次排名':
                    cell.fill = light_pink_purple_fill
                # 进退步用紫色
                elif col_name and isinstance(col_name, str) and col_name == '进退步':
                    cell.fill = purple_fill
                else:
                    cell.fill = highlight_fill if is_highlight else row_fill
                
                cell.alignment = data_alignment
                cell.border = border
            except AttributeError:
                # 忽略合并单元格的只读属性错误
                pass

def _format_class_sheet(worksheet, df):
    """设置班级汇总表的格式：表头颜色、奇偶行颜色、边框、分数段上色"""
    from config import COLOR_CLASS as color, COLOR_CLASS_SCORE_SEGMENT

    header_fill = _get_fill(color['header_bg'])
    header_font = Font(bold=True, color=color['header_fg'][1:])
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = _get_border(color['border'])

    # 表头现在在第1行（因为不再插入标题行）
    header_row = 1
    
    # 格式化表头行
    headers = []
    for col_idx in range(1, worksheet.max_column + 1):
        try:
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
            headers.append(cell.value)
        except AttributeError:
            # 忽略合并单元格的只读属性错误
            headers.append(None)

    odd_fill = _get_fill(color['odd_row_bg'])
    even_fill = _get_fill(color['even_row_bg'])
    data_alignment = Alignment(horizontal='center', vertical='center')

    # 识别分数段列
    score_segment_columns = _get_score_segment_columns(headers)
    
    # 区分班级区和年段区
    class_max_value = 0  # 班级区最大值
    year_max_value = 0   # 年段区最大值
    class_rows = []      # 班级行索引
    year_rows = []       # 年段行索引
    
    if score_segment_columns:
        data_start_row = header_row + 1
        
        # 首先识别班级行和年段行
        class_column_idx = None
        # 查找班级列
        for i, header in enumerate(headers):
            if header and str(header).strip() == '班级':
                class_column_idx = i + 1  # 转换为1-based索引
                break
        
        # 收集班级行和年段行
        for row_idx in range(data_start_row, worksheet.max_row + 1):
            is_class_row = True
            if class_column_idx:
                try:
                    cell = worksheet.cell(row=row_idx, column=class_column_idx)
                    if cell.value:
                        class_name = str(cell.value).strip()
                        # 识别年段行（根据班级名称判断）
                        if class_name in ['年段', '全校', '全年级', '年级', '总', '总计', '合计']:
                            is_class_row = False
                            year_rows.append(row_idx)
                        else:
                            class_rows.append(row_idx)
                    else:
                        class_rows.append(row_idx)
                except AttributeError:
                    class_rows.append(row_idx)
            else:
                # 如果找不到班级列，默认所有行都是班级行
                class_rows.append(row_idx)
        
        # 计算班级区的最大值
        class_score_values = []
        for row_idx in class_rows:
            for col_idx in score_segment_columns:
                try:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        try:
                            val = float(cell.value)
                            class_score_values.append(val)
                        except (ValueError, TypeError):
                            pass
                except AttributeError:
                    pass
        if class_score_values:
            class_max_value = max(class_score_values)
        
        # 计算年段区的最大值
        year_score_values = []
        for row_idx in year_rows:
            for col_idx in score_segment_columns:
                try:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        try:
                            val = float(cell.value)
                            year_score_values.append(val)
                        except (ValueError, TypeError):
                            pass
                except AttributeError:
                    pass
        if year_score_values:
            year_max_value = max(year_score_values)
        
        # 如果班级区或年段区没有值，使用另一个区的最大值作为备用
        if class_max_value == 0 and year_max_value > 0:
            class_max_value = year_max_value
        elif year_max_value == 0 and class_max_value > 0:
            year_max_value = class_max_value

    # 格式化数据行
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        row_fill = even_fill if (row_idx - header_row - 1) % 2 == 0 else odd_fill
        for col_idx in range(1, worksheet.max_column + 1):
            try:
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # 检查是否是分数段列
                if col_idx in score_segment_columns and cell.value is not None:
                    # 对分数段列应用上色
                    try:
                        value = float(cell.value)
                        # 根据行类型选择对应的最大值
                        if row_idx in class_rows:
                            # 班级行使用班级区的最大值
                            cell.fill = _get_score_segment_color(value, class_max_value, COLOR_CLASS_SCORE_SEGMENT)
                        elif row_idx in year_rows:
                            # 年段行使用年段区的最大值
                            cell.fill = _get_score_segment_color(value, year_max_value, COLOR_CLASS_SCORE_SEGMENT)
                        else:
                            # 其他行使用默认填充
                            cell.fill = row_fill
                    except (ValueError, TypeError):
                        # 如果不是数值，使用默认填充
                        cell.fill = row_fill
                else:
                    # 非分数段列使用默认填充
                    cell.fill = row_fill
                
                cell.alignment = data_alignment
                cell.border = border
            except AttributeError:
                # 忽略合并单元格的只读属性错误
                pass

def _adjust_column_width(worksheet):
    """根据内容自动调整列宽"""
    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                # 检查单元格是否是合并单元格的一部分
                # 只有左上角的单元格可以访问value属性
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except AttributeError:
                # 忽略合并单元格的只读属性错误
                pass
        adjusted_width = min(max(max_length + 2, 8), 40)
        worksheet.column_dimensions[col_letter].width = adjusted_width